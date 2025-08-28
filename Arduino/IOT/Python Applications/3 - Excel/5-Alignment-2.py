# Cell Alignment Setting

from openpyxl import Workbook
from openpyxl.styles import Alignment  
import time

wb = Workbook()
ws = wb.active  # reference to the active sheet

ws.row_dimensions[1].height = 20
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 15

ws['A1'] = "Name"
ws['B1'] = "Age"
ws['C1'] = "Course"
ws['D1'] = "College Name"
ws['E1'] = "Mobile Number"
ws['F1'] = "Percentage"

currentCell = ws['A1']       
currentCell.alignment = Alignment(horizontal='center', vertical='center')  # use "right", "left"

ws['A2'] = input("Enter the Name ")
ws['B2'] = input("Enter the Age ")
ws['C2'] = input("Enter the Course ")
ws['D2'] = input("Enter the College Name ")
ws['E2'] = input("Enter the Mobile Number ")
ws['F2'] = float(input("Enter the Percentage "))


wb.save("College.xlsx")
