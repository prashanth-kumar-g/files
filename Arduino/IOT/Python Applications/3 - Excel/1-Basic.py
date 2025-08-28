from openpyxl import Workbook
import time

wb = Workbook()
ws = wb.active  # reference to the active sheet

ws['A1'] = 10
ws['A2'] = 20
ws['A3'] = 30

ws['B1'] = 56
ws['B2'] = 43
ws['B3'] = 77

wb.save("College.xlsx")
