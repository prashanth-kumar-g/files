# Width and Height Setting

from openpyxl import Workbook
import time

wb = Workbook()
ws = wb.active  # reference to the active sheet

ws.row_dimensions[1].height = 20
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 15

ws['A1'] = "Name"
ws['B1'] = "Age"
ws['C1'] = "Course"
ws['D1'] = "College Name"
ws['E1'] = "Mobile Number"
ws['F1'] = "Percentage"


while True:
    a = input("Enter the Name ")
    b = input("Enter the Age ")
    c = input("Enter the Course ")
    d = input("Enter the College Name ")
    e = input("Enter the Mobile Number ")
    f = float(input("Enter the Percentage "))


    # Appendin the Values to Excel Sheet
    ws.append([a,b,c,d,e,f])

    # Saving the  Excel Sheet
    wb.save("College.xlsx")

    
    print("Contents Saved in the Excel Sheet ")
    print("Do you wnat to Continue ? ")
    ch = int(input("Press 1 to Continue or Other Key to Exit "))
    if ch !=1:
        break
        
    
