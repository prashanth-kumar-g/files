from cvzone.SerialModule import SerialObject
import time
import sys
import datetime
from openpyxl import Workbook


wb = Workbook()
ws = wb.active

arduino = SerialObject('/dev/ttyUSB0')

ws.row_dimensions[1].height = 40
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15

ws['A1'] = "Date"
ws['B1'] = "Time"
ws['C1'] = "Temperature"
ws['D1'] = "Humidity"


while True:
    Data = arduino.getData()
    print("The Temperature == ", Data[0])
    print("The Humidity == ", Data[1])
    ws.append([time.strftime('%d-%m-%Y'), time.strftime("%H:%M:%S"), Data[0], Data[1]])
    wb.save('Weather-Report.xls')
    

