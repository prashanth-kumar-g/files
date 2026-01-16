# Width and Height Setting

from openpyxl import Workbook
import time

wb = Workbook()
ws = wb.active  # reference to the active sheet

ws.row_dimensions[1].height = 20
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 15

ws['A1'] = "Name"
ws['B1'] = "Age"
ws['C1'] = "Course"
ws['D1'] = "College Name"
ws['E1'] = "Mobile Number"
ws['F1'] = "Percentage"

wb.save("College.xlsx")
